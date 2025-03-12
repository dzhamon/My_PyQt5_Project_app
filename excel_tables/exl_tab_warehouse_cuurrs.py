import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def excel_table_warehouses_currs(df):
    """
        Метод вывода датафрейма result_df в файл Отчет по остаткам на складах в валютах котракта
        warehouse_totals_formatted.xlsx.xlsx
        :param df:
        :return:
    """
    # Создаем новый Excel-файл
    output_file = "warehouse_totals_formatted.xlsx"
    wb = Workbook()
    ws = wb.active
    
    # Заголовок таблицы
    ws.title = "Warehouse Totals"
    ws['A1'] = "Отчет по остаткам на складах"
    ws['A1'].font = Font(name='Arial Narrow', size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Создаем стиль для числовых значений
    number_style = NamedStyle(name="number_style", number_format='#,##0.00')
    wb.add_named_style(number_style)
    
    # Добавляем DataFrame в Excel, начиная с 3-й строки (чтобы оставить место для заголовка)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(name='Arial Narrow', size=12)  # Шрифт для данных
            if r_idx == 3:  # Заголовки столбцов
                cell.font = Font(name='Arial Narrow', size=12, bold=True)
                cell.alignment = Alignment(horizontal='center')
            if isinstance(value, (int, float)): # применяем числовой формат к числам
                cell.style = number_style
    
    # Настраиваем ширину столбцов
    column_widths = {
        'A': 15,  # Ширина столбца для warehouse
        'B': 15,  # Ширина столбца для currency
        'C': 20   # Ширина столбца для total_price_currency
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Сохраняем файл
    wb.save(output_file)
    
    # Открываем файл (для просмотра)
    os.startfile(output_file)  # Для Windows
    
    print(f"Файл сохранен: {output_file}")